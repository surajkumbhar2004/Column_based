from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from rapidfuzz import fuzz, process
import pandas as pd
from io import BytesIO
import tempfile
import re

app = FastAPI()

# Allow CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/", response_class=HTMLResponse)
async def root():
    return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>File Comparison Tool</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #eef2f5;
            margin: 0;
            padding: 20px;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }
        .container {
            width: 100%;
            max-width: 800px;
            background: #ffffff;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 12px 30px rgba(0, 0, 0, 0.1);
        }
        h1 {
            color: #333;
            font-size: 28px;
            text-align: center;
            margin-bottom: 20px;
        }
        .form-row {
            display: flex;
            justify-content: space-between;
            margin-bottom: 20px;
        }
        .form-group {
            width: 48%;
        }
        label {
            display: block;
            font-weight: 600;
            margin-bottom: 8px;
            color: #555;
        }
        input[type="file"], select {
            width: 100%;
            padding: 10px;
            border: 2px solid #ccc;
            border-radius: 8px;
            box-sizing: border-box;
            transition: border-color 0.3s ease;
        }
        input[type="file"]:hover, select:hover {
            border-color: #4fa3f7;
        }
        button {
            width: 100%;
            padding: 15px;
            background-color: #4fa3f7;
            color: white;
            font-size: 18px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            margin-top: 10px;
            transition: background-color 0.3s ease;
        }
        button:hover {
            background-color: #3578d3;
        }
        #loading {
            display: none;
            margin-top: 20px;
            text-align: center;
            color: #333;
        }
        .spinner {
            border: 6px solid #f3f3f3;
            border-top: 6px solid #4fa3f7;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: auto;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        #errorContainer {
            color: #d9534f;
            margin-top: 10px;
            font-weight: 600;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Upload Files</h1>
        <form id="uploadForm" enctype="multipart/form-data">
            <div class="form-row">
                <div class="form-group">
                    <label for="file1">Source File:</label>
                    <input type="file" id="file1" name="file1" required onchange="handleFileSelect()">
                </div>
                <div class="form-group">
                    <label for="sourceCol">Select Source Column:</label>
                    <select id="sourceCol"></select>
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label for="file2">Target File:</label>
                    <input type="file" id="file2" name="file2" required onchange="handleFileSelect()">
                </div>
                <div class="form-group">
                    <label for="targetCol">Select Target Column:</label>
                    <select id="targetCol"></select>
                </div>
            </div>
            <button type="button" id="compareBtn" onclick="compareFiles()">Compare Files</button>
            <div id="loading">
                <p>Processing...</p>
                <div class="spinner"></div>
            </div>
            <div id="errorContainer"></div>
        </form>
    </div>

<script>
async function handleFileSelect() {
    const file1 = document.getElementById('file1').files[0];
    const file2 = document.getElementById('file2').files[0];
    
    if (file1 && file2) {
        const formData = new FormData();
        formData.append('file1', file1);
        formData.append('file2', file2);
        
        try {
            const response = await fetch("/get_columns/", { method: "POST", body: formData });
            if (!response.ok) {
                const error = await response.json();
                document.getElementById("errorContainer").innerText = error.detail || "Error reading columns. Please check the files.";
                return;
            }

            const columns = await response.json();
            populateDropdown('sourceCol', columns.file1);
            populateDropdown('targetCol', columns.file2);
        } catch (error) {
            document.getElementById("errorContainer").innerText = "Error fetching column names.";
        }
    }
}

function populateDropdown(dropdownId, columns) {
    const dropdown = document.getElementById(dropdownId);
    dropdown.innerHTML = columns.map(col => `<option value="${col}">${col}</option>`).join('');
}

async function compareFiles() {
    const sourceCol = document.getElementById('sourceCol').value;
    const targetCol = document.getElementById('targetCol').value;
    
    if (!sourceCol || !targetCol) return;

    document.getElementById("loading").style.display = 'block';
    document.getElementById("errorContainer").innerText = "";

    const formData = new FormData();
    formData.append('file1', document.getElementById('file1').files[0]);
    formData.append('file2', document.getElementById('file2').files[0]);
    formData.append('sourceCol', sourceCol);
    formData.append('targetCol', targetCol);

    try {
        const response = await fetch("/compare/", { method: "POST", body: formData });
        document.getElementById("loading").style.display = 'none';

        if (!response.ok) {
            const result = await response.json();
            document.getElementById("errorContainer").innerText = result.error || "Error processing files.";
            return;
        }

        const blob = await response.blob();
        const url = window.URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "output.xlsx";
        document.body.appendChild(a);
        a.click();
    } catch (error) {
        document.getElementById("errorContainer").innerText = "An error occurred during comparison.";
    }
}
</script>
</body>
</html>
"""

@app.post("/get_columns/")
async def get_columns(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    try:
        # Read the files into DataFrames
        df1 = pd.read_excel(BytesIO(await file1.read()))
        df2 = pd.read_excel(BytesIO(await file2.read()))
        
        # Return column names
        return {"file1": df1.columns.tolist(), "file2": df2.columns.tolist()}
    except ValueError as ve:
        # Likely an issue with file format or parsing
        raise HTTPException(status_code=400, detail=f"Error reading files: {str(ve)}. Ensure files are valid Excel files.")
    except Exception as e:
        # General exception handling
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")


@app.post("/compare/")
async def compare_files(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    sourceCol: str = Form(...),
    targetCol: str = Form(...),
    threshold: int = 75
):
    try:
        # Load Excel files into DataFrames
        df1 = pd.read_excel(BytesIO(await file1.read()))
        df2 = pd.read_excel(BytesIO(await file2.read()))

        source_dtype = df1[sourceCol].dtype
        target_dtype = df2[targetCol].dtype
        
        # Enhanced data type check
        if pd.api.types.is_numeric_dtype(df1[sourceCol]) and pd.api.types.is_numeric_dtype(df2[targetCol]):
            pass  # Numeric types, exact match
        elif pd.api.types.is_string_dtype(df1[sourceCol]) and pd.api.types.is_string_dtype(df2[targetCol]):
            pass  # String types, exact match for strings
        else:
            # If types are not matching
            return JSONResponse(status_code=400, content={"error": "Datatypes are not similar. Please select similar data types."})

        match_status, match_scores, matched_values = [], [], []

        if pd.api.types.is_numeric_dtype(df1[sourceCol]):
            # Numeric Columns: Exact match only
            for val1 in df1[sourceCol]:
                if val1 in df2[targetCol].values:
                    match_status.append("Exact Match")
                    match_scores.append(100)
                    matched_values.append(val1)  # Assign the exact match
                else:
                    match_status.append("No Match")
                    match_scores.append(0)
                    matched_values.append(None)  # Leave empty if no match
        else:
            # String Columns: Clean and compare
            def clean_string(value):
                """Clean string by removing unwanted characters."""
                if pd.isna(value):
                    return ''
                return re.sub(r'[\\/\.]', '', str(value).strip())

            df1[sourceCol] = df1[sourceCol].apply(clean_string)
            df2[targetCol] = df2[targetCol].apply(clean_string)

            for val1 in df1[sourceCol]:
                if val1 in df2[targetCol].values:
                    # Exact match
                    match_status.append("Exact Match")
                    match_scores.append(100)
                    matched_values.append(val1)
                else:
                    # Fuzzy match
                 close_matches = process.extract(val1, df2[targetCol].dropna().tolist(), scorer=fuzz.token_sort_ratio, limit=1)
                 if close_matches:
                    best_match, best_score, _ = close_matches[0]  # Handle the additional index returned by `process.extract`
                 else:
                     best_match, best_score = None, 0

                 if best_score >= threshold:
                    match_status.append("Close Match")
                    match_scores.append(best_score)
                    matched_values.append(best_match)  # Assign fuzzy match if above threshold
                 else:
                    match_status.append("No Match")
                    match_scores.append(0)
                    matched_values.append(None)  # Leave empty if no match


        # Add match results to the DataFrame
        df1["Match Status"] = match_status
        df1["Match Score"] = match_scores
        df1["Target Column"] = matched_values

        # Define color fills
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active

        # Write data and apply styles
        for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
                if r_idx > 1:  # Skip the header row
                    # Apply color fill to the selected source column
                    if c_idx == list(df1.columns).index(sourceCol) + 1:  # Check if it's the selected source column
                        match_status = row[list(df1.columns).index("Match Status")]
                
                        # Apply color based on match status
                        if match_status == "Exact Match":
                            cell.fill = green_fill
                        elif match_status == "Close Match":
                            cell.fill = orange_fill
                        elif match_status == "No Match":
                            cell.fill = red_fill

        # Save result to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name

        return FileResponse(tmp_file_path, headers={"Content-Disposition": "attachment; filename=output.xlsx"})
    
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"An error occurred: {str(e)}"})


if __name__ == "__main__":
    import uvicorn
 
    uvicorn.run(app, host="127.0.0.1", port=8080)